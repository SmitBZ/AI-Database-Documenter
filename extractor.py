# -*- coding: utf-8 -*-
"""
Nombre del Proyecto: AI Database Documenter
Descripción: Script para automatizar la generación de diccionarios de datos SQL.
Autor: Smit BZ
Fecha de creación: 2026
Licencia: MIT
"""
# Librerias para el funcionamiento
import re
import json
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import create_engine, inspect, text
import customtkinter as ctk
from groq import Groq

# Colores que se usaran para las tablas
COLOR_HEADER_BG   = "1F3864"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_LABEL_BG    = "D9E1F2"
COLOR_PK_BG       = "FFF2CC"
COLOR_PK_TEXT     = "7D4A00"
COLOR_FK_BG       = "E2EFDA"
COLOR_FK_TEXT     = "297238"
COLOR_PKFK_BG     = "FCE4D6"
COLOR_PKFK_TEXT   = "843C0C"

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# Campos pre establecidos para llenar.
"""
    Para este campo, se puede llenar algunas constantes o variables de los atributos.
    Util si vas a documentar multiples esquemas de BD.
"""
# Variables de ejemplo
FIXED_MAPS = {
    "pers_id": ("ID Persona", "Identificador de persona."),
    "pers_nombre": ("Nombre", "Nombre de persona. Ej. Carlos Ismael"),
    "ciud_id": ("ID Ciudad", "Identificador de ciudad"),
    "audit_user_pc":      ("Usuario Windows",  "Nombre de usuario de la sesión de PC."),
    "audit_nombre_pc":    ("Nombre del Host",  "Nombre del equipo desde donde se registró."),
    "audit_ip":           ("Dirección IP",     "IP del terminal del operador."),
    "audit_mac_address":  ("Dirección MAC",    "Dirección física de la tarjeta de red."),
    "created_at": ("Fecha Creación", "Fecha y hora de creación del registro. DEFAULT NOW()."),
    "updated_at": ("Fecha Actualización", "Fecha y hora de actualización de un registro. DEFAULT NOW()."),
    "usua_id": ("ID de Usuario", "Usuario que ejecutó la operación."),
    "audit_tb_afectada": ("Tabla afectada", "Nombre de la tabla que fue afectada por la operación realizada. Ej.: "),
    "audit_operacion": ("Tipo de operación", "INSERT, UPDATE, DELETE. CHECK constraint."),
    "audit_usuario_id": ("ID de usuario", "Identificador del usuario que realizó la operación."),
    "audit_cambio_ant": ("Valor anterior", "Valor previo a la modificación en formato texto. NULL en INSERT."),
    "audit_cambio_actual": ("Valor actual", "Valor nuevo después de la modificación. NULL en DELETE."),
    "audit_descripcion": ("Descripción de la operación", "Almacena una descripción detallada de la operación realizada."),
    "audit_fecha_operacion": ("Fecha de la operación", "Fecha y hora en que se realizó la operación. DEFAULT NOW()."),
    "audit_ip_psql": ("IP de PostgreSQL", "Dirección IP del servidor de PostgreSQL en el que se realizó la operación."),
    "audit_os": ("Sistema operativo", "Sistema operativo de la computadora desde la que se realizó la operación."),
    "acta_id": ("ID Acta", "Identificador del acta."),
    "depe_id" : ("ID dependencia", "Identificador de la dependencia."),
}
# Configuración por defecto para cada motor de BD soportado.
# Incluye puerto, usuario y cadena de driver para SQLAlchemy.
DB_DEFAULTS = {
    "PostgreSQL":           {"port": "5432",  "user": "postgres", "driver": "postgresql+psycopg2"},
    "MySQL":                {"port": "3306",  "user": "root",     "driver": "mysql+pymysql"},
    "MariaDB":              {"port": "3306",  "user": "root",     "driver": "mysql+pymysql"},
    "Oracle Database":      {"port": "1521",  "user": "system",   "driver": "oracle+cx_oracle"},
}


# Petición para el llenado por la IA
class GroqEnricher:
    """
    Realiza UNA sola llamada por tabla a la API de Groq y devuelve:
      - descripcion   : descripción funcional de la tabla
      - claves_unicas : columnas únicas aparte de la PK
      - observaciones : notas técnicas o de negocio
      - columns       : {col_name: {logico, descripcion}}  (para cada columna)

    Usa el SDK oficial `groq` con response_format=json_object para
    obtener JSON puro sin necesidad de limpiar markdown.
    """
    # En futuros modelos, se cambia por el modelo para seguir usando
    MODEL = "llama-3.3-70b-versatile"

    def __init__(self, api_key: str):
        self.client = Groq(api_key=api_key.strip())

    def enrich_table(self, t_name: str, columns: list,
                     pk_cols: list, fk_map: dict,
                     schema: str | None) -> dict:
        col_lines = []
        for col in columns:
            roles = []
            if col["name"] in pk_cols:
                roles.append("PK")
            if col["name"] in fk_map:
                roles.append(f"FK→{fk_map[col['name']]}")
            nullable = "NULL" if col.get("nullable", True) else "NOT NULL"
            role_str = f", {'/'.join(roles)}" if roles else ""
            col_lines.append(
                f"  - {col['name']} ({col.get('type', '?')}, {nullable}{role_str})"
            )
        pk_str  = ", ".join(pk_cols) if pk_cols else "ninguna"
        fk_str  = ", ".join(f"{k}→{v}" for k, v in fk_map.items()) if fk_map else "ninguna"
        sch_str = schema or "público"

        system_msg = (
            "Eres un experto en modelado de bases de datos relacionales y documentación "
            "técnica en español. Responde SIEMPRE con un objeto JSON válido y completo, "
            "sin texto adicional ni bloques markdown. "
            "Reglas OBLIGATORIAS:\n"
            "1. Nunca uses la palabra 'Almacena'.\n"
            "2. Sé concreto, técnico y sencillo.\n"
            "3. En la descripción de CADA campo incluye siempre un ejemplo real con el formato: Ej.: <valor>.\n"
            "4. Para campos con restricciones CHECK, menciónalas explícitamente (ej.: FIJO, ROTATIVO, ESPECIAL).\n"
            "5. Para campos FK indica la tabla referenciada (ej.: FK → TABLA.columna).\n"
            "6. Si no tienes información suficiente para alguna sección, déjala vacía pero mantén la estructura del JSON."
        )
        user_msg = (
            f"Analiza esta tabla de base de datos y genera su documentación en español.\n\n"
            f"Tabla   : {t_name}\n"
            f"Esquema : {sch_str}\n"
            f"PK      : {pk_str}\n"
            f"FK      : {fk_str}\n"
            f"Columnas:\n" + "\n".join(col_lines) + "\n\n"
            "Devuelve EXACTAMENTE este JSON (sin claves adicionales):\n"
            "{\n"
            '  "descripcion":   "Descripción funcional de la tabla (1-2 oraciones). Sin ejemplos aquí.",\n'
            '  "claves_unicas": "Columnas únicas aparte de la PK (ej.: col_codigo), o \\"Ninguna\\".",\n'
            '  "observaciones": "Notas técnicas o de negocio relevantes para el uso de esta tabla.",\n'
            '  "columns": {\n'
            '    "<nombre_columna>": {\n'
            '      "logico":      "Nombre legible en español (2-4 palabras)",\n'
            '      "descripcion": "Qué representa este campo. Restricciones si aplica. Ej.: <valor_ejemplo>"\n'
            "    }\n"
            "  }\n"
            "}\n"
            "Incluye TODAS las columnas listadas en el bloque 'columns'. "
            "Recuerda: CADA descripción de columna DEBE terminar con un ejemplo usando el formato 'Ej.: <valor>'."
        )
        completion = self.client.chat.completions.create(
            model=self.MODEL,
            temperature=0.3,
            max_tokens=2048,
            response_format={"type": "json_object"},   # garantiza JSON puro
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = completion.choices[0].message.content
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return {
                "descripcion": "", "claves_unicas": "",
                "observaciones": "", "columns": {},
            }
    def generate_relations_and_indexes(self, tables_data: dict) -> dict:
        """
        Genera en una sola llamada las relaciones entre tablas y los índices
        de la base de datos a partir de toda la estructura conocida.
        Devuelve: {"relaciones": [...], "indices": [...]}
        """
        summary_lines = []
        for t_name, info in tables_data.items():
            pk_str = ", ".join(info.get("pk", [])) or "ninguna"
            fk_items = ", ".join(f"{k}→{v}" for k, v in info.get("fk_map", {}).items()) or "ninguna"
            cols = ", ".join(c["name"] for c in info.get("columns", []))
            summary_lines.append(
                f"Tabla: {t_name} | PK: {pk_str} | FK: {fk_items} | Columnas: {cols}"
            )
        system_msg = (
            "Eres un experto en modelado de bases de datos relacionales. "
            "Responde SIEMPRE con un objeto JSON válido y completo, sin texto adicional ni bloques markdown."
        )
        user_msg = (
            "Analiza el siguiente esquema de base de datos y genera:\n"
            "1. Las relaciones entre tablas (basadas en las FK detectadas).\n"
            "2. Los índices recomendados para optimizar consultas.\n\n"
            "Esquema:\n" + "\n".join(summary_lines) + "\n\n"
            "Devuelve EXACTAMENTE este JSON:\n"
            "{\n"
            '  "relaciones": [\n'
            '    {\n'
            '      "tabla_origen": "NOMBRE_TABLA",\n'
            '      "tabla_destino": "NOMBRE_TABLA",\n'
            '      "tipo_relacion": "1 : N",\n'
            '      "descripcion": "Descripción de la relación incluyendo FK y ON DELETE."\n'
            '    }\n'
            '  ],\n'
            '  "indices": [\n'
            '    {\n'
            '      "nombre_indice": "pk_tabla",\n'
            '      "tabla": "NOMBRE_TABLA",\n'
            '      "tipo": "B-Tree",\n'
            '      "columnas": "col1, col2",\n'
            '      "unico": "Sí"\n'
            '    }\n'
            '  ]\n'
            "}\n"
            "Tipos de relación válidos: 1:1, 1:N, N:1, N:M. "
            "Para índices incluye PKs (únicos), FKs (no únicos) y columnas de búsqueda frecuente."
        )
        completion = self.client.chat.completions.create(
            model=self.MODEL,
            temperature=0.2,
            max_tokens=4096,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = completion.choices[0].message.content
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return {"relaciones": [], "indices": []}


# Sección por si deseas documentar archivos sql
class SQLFileParser:
    """
    Extrae estructura DDL desde archivos .sql / backups.
    Compatible con PostgreSQL, MySQL y MariaDB.

    Métodos principales:
        parse(filepath)  : Lee el archivo y devuelve un dict {tabla: {columns, pk, fk_map}}.
        _parse_body(body): Analiza el cuerpo de un CREATE TABLE.
        _parse_col(defn) : Extrae nombre, tipo, longitud y restricciones de una columna.
        _split_defs(body): Divide las definiciones de columnas respetando paréntesis anidados.
    """
    _COMMENT_SINGLE = re.compile(r'--[^\n]*')
    _COMMENT_MULTI  = re.compile(r'/\*.*?\*/', re.DOTALL)

    def parse(self, filepath: str) -> dict:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            sql = f.read()
        sql = self._COMMENT_SINGLE.sub("", sql)
        sql = self._COMMENT_MULTI.sub("", sql)
        tables = {}
        for m in re.finditer(
            r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?'
            r'(?:`?[\w$]+`?\.)?`?([\w$]+)`?\s*\((.+?)\)\s*'
            r'(?:ENGINE|TABLESPACE|;|\Z)',
            sql, re.IGNORECASE | re.DOTALL,
        ):
            tables[m.group(1)] = self._parse_body(m.group(2))
        return tables

    def _parse_body(self, body: str) -> dict:
        pk_cols, fk_map, columns = [], {}, []
        for defn in self._split_defs(body):
            defn = defn.strip()
            if not defn:
                continue
            upper = defn.upper().lstrip()

            if upper.startswith("PRIMARY KEY"):
                pk_cols.extend(re.findall(r'`?([\w$]+)`?', defn.split("(", 1)[-1]))

            elif upper.startswith(("FOREIGN KEY", "CONSTRAINT")):
                fkm = re.search(
                    r'FOREIGN\s+KEY\s*\(([^)]+)\)\s*REFERENCES\s+`?([\w$]+)`?\s*\(([^)]+)\)',
                    defn, re.IGNORECASE,
                )
                if fkm:
                    srcs = [c.strip().strip("`") for c in fkm.group(1).split(",")]
                    refs = [c.strip().strip("`") for c in fkm.group(3).split(",")]
                    for sc, rc in zip(srcs, refs):
                        fk_map[sc] = f"{fkm.group(2).strip()}.{rc}"

            elif not upper.startswith(("UNIQUE", "INDEX", "KEY ", "CHECK")):
                col = self._parse_col(defn)
                if col:
                    columns.append(col)

        for col in columns:
            if col.get("inline_pk"):
                pk_cols.append(col["name"])
        return {"columns": columns, "pk": pk_cols, "fk_map": fk_map}

    def _parse_col(self, defn: str) -> dict | None:
        m = re.match(r'`?([\w$]+)`?\s+([\w]+(?:\s*\([^)]*\))?)', defn, re.IGNORECASE)
        if not m:
            return None
        type_raw  = m.group(2)
        len_match = re.search(r'\((\d+)(?:,\d+)?\)', type_raw)
        return {
            "name":      m.group(1),
            "type":      type_raw.split("(")[0].upper(),
            "length":    len_match.group(1) if len_match else "—",
            "nullable":  "NOT NULL" not in defn.upper(),
            "inline_pk": "PRIMARY KEY" in defn.upper(),
        }

    def _split_defs(self, body: str) -> list:
        parts, depth, cur = [], 0, []
        for ch in body:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            if ch == "," and depth == 0:
                parts.append("".join(cur))
                cur = []
            else:
                cur.append(ch)
        if cur:
            parts.append("".join(cur))
        return parts

# Esta es la seccion donde se genera el excel
class ExcelGenerator:
    """
    Genera y formatea el archivo Excel del diccionario de datos.

    Responsabilidades:
        - Aplicar estilos (colores, fuentes, bordes) a las celdas según su rol (PK, FK, header, etc.).
        - Determinar el nombre lógico y descripción de cada columna usando FIXED_MAPS o la IA.
        - Escribir una hoja por tabla con sus metadatos y columnas.
        - Escribir la hoja de relaciones entre tablas.
        - Escribir la hoja de índices recomendados.
    """
    def apply_style(self, cell, style_type="data",
                    is_pk=False, is_fk=False, is_pkfk=False):
        """
        Aplica formato visual a una celda de Excel.

        Args:
            cell       : Celda de openpyxl a formatear.
            style_type : "header" | "label" | "data" (por defecto).
            is_pk      : True si la columna es clave primaria.
            is_fk      : True si la columna es clave foránea.
            is_pkfk    : True si la columna es PK y FK simultáneamente.
        """
        cell.font      = Font(name="Calibri", size=9)
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        if style_type == "header":
            cell.fill  = PatternFill(start_color=COLOR_HEADER_BG,
                                     end_color=COLOR_HEADER_BG, fill_type="solid")
            cell.font  = Font(name="Arial", size=9, bold=True, color=COLOR_HEADER_TEXT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif style_type == "label":
            cell.fill = PatternFill(start_color=COLOR_LABEL_BG,
                                    end_color=COLOR_LABEL_BG, fill_type="solid")
            cell.font = Font(name="Arial", size=9, bold=True)
        elif is_pkfk:
            cell.fill = PatternFill(start_color=COLOR_PKFK_BG,
                                    end_color=COLOR_PKFK_BG, fill_type="solid")
            cell.font = Font(name="Arial", size=9, bold=True, color=COLOR_PKFK_TEXT)
        elif is_pk:
            cell.fill = PatternFill(start_color=COLOR_PK_BG,
                                    end_color=COLOR_PK_BG, fill_type="solid")
            cell.font = Font(name="Arial", size=9, bold=True, color=COLOR_PK_TEXT)
        elif is_fk:
            cell.fill = PatternFill(start_color=COLOR_FK_BG,
                                    end_color=COLOR_FK_BG, fill_type="solid")
            cell.font = Font(name="Arial", size=9, color=COLOR_FK_TEXT)

    # ── Relleno automático de nombre lógico y descripción ────────
    def get_auto_fill(self, col_name: str, table_name: str,
                      is_pk: bool, fk_map: dict) -> tuple:
        """
        Retorna (nombre_lógico, descripción) para una columna, priorizando:
        1. FIXED_MAPS  → mapeo estático pre-definido por el usuario.
        2. fk_map      → nombre inferido a partir de la tabla referenciada.
        3. PK genérica → "ID <NombreTabla>".
        4. ("", "")    → sin información disponible (la IA intentará completarlo).

        Args:
            col_name   : Nombre físico de la columna.
            table_name : Nombre físico de la tabla.
            is_pk      : True si la columna es PK.
            fk_map     : Diccionario {col: "tabla.columna"} de claves foráneas.

        Returns:
            Tupla (nombre_lógico: str, descripcion: str).
        """
        if col_name in FIXED_MAPS:
            return FIXED_MAPS[col_name]
        if col_name in fk_map:
            ref      = fk_map[col_name]
            ref_name = ref.split(".")[0].replace("tb_", "").replace("_", " ").title()
            label    = "PK/FK" if is_pk else "FK"
            return f"ID {ref_name}", f"{label} → {ref}"
        if is_pk:
            clean = table_name.replace("tb_", "").replace("_", " ").title()
            return f"ID {clean}", "Identificador único incremental."
        return ("", "")

    # ── Conversión de tipo para llaves primarias ──────────────────
    @staticmethod
    def _pk_display_type(tipo: str) -> str:
        """Convierte INTEGER→SERIAL y BIGINT→BIGSERIAL solo para PKs."""
        t = tipo.upper()
        if t == "INTEGER":
            return "SERIAL"
        if t == "BIGINT":
            return "BIGSERIAL"
        return tipo

    # ── Escritura de hoja por tabla ───────────────────────────────
    def write_table_sheet(self, wb, t_name: str, columns: list,
                          pk_cols: list, fk_map: dict,
                          schema: str | None = None,
                          enriched: dict | None = None):
        ws = wb.create_sheet(title=t_name[:30].upper())

        # Relaciones desde FK map
        rel_text = ", ".join(
            f"{v.split('.')[0].upper()} (N:1)" for v in fk_map.values()
        )
        physical_name = f"{schema}.{t_name}" if schema else t_name

        # Clave Primaria con tipo de dato incluido
        pk_parts = []
        for pk_col in pk_cols:
            col_info = next((c for c in columns if c["name"] == pk_col), None)
            tipo = col_info.get("type", "").upper() if col_info else ""
            tipo_display = self._pk_display_type(tipo) if tipo else ""
            pk_parts.append(
                f"{pk_col} ({tipo_display})" if tipo_display else pk_col
            )
        pk_text = ", ".join(pk_parts)

        # Campos enriquecidos por IA (vacío si no hay)
        ai            = enriched or {}
        descripcion   = ai.get("descripcion",   "") or ""
        claves_unicas = ai.get("claves_unicas",  "") or ""
        observaciones = ai.get("observaciones",  "") or ""

        # ── Bloque de metadatos ───────────────────────────────────
        meta = [
            ["Nombre Físico",  physical_name],
            ["Descripción",    descripcion],
            ["Clave Primaria", pk_text],
            ["Claves Únicas",  claves_unicas],
            ["Relaciones",     rel_text],
            ["Observaciones",  observaciones],
        ]
        for row in meta:
            ws.append(row)
            r = ws.max_row
            self.apply_style(ws.cell(row=r, column=1), "label")
            self.apply_style(ws.cell(row=r, column=2))
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)

        # ── Cabecera de columnas ──────────────────────────────────
        ws.append([])
        headers = ["Campo (Físico)", "Nombre Lógico", "Tipo Dato",
                   "Long.", "Nulo", "Clave", "Descripción"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            self.apply_style(cell, "header")

        # ── Filas de columnas ─────────────────────────────────────
        ai_cols = ai.get("columns", {})
        for col in columns:
            c_name  = col["name"]
            is_pk   = c_name in pk_cols
            is_fk   = c_name in fk_map
            is_pkfk = is_pk and is_fk
            key_lbl = "PK/FK" if is_pkfk else ("PK" if is_pk else ("FK" if is_fk else ""))

            logico, desc = self.get_auto_fill(c_name, t_name, is_pk, fk_map)

            # Completar con IA solo si el campo sigue vacío
            ai_col = ai_cols.get(c_name, {})
            if not logico:
                logico = ai_col.get("logico", "")
            if not desc:
                desc = ai_col.get("descripcion", "")

            col_type = col.get("type", "")
            if is_pk:
                col_type = self._pk_display_type(col_type)
            ws.append([
                c_name,
                logico,
                col_type,
                col.get("length", "—"),
                "N" if not col.get("nullable", True) else "S",
                key_lbl,
                desc,
            ])
            for c_idx in range(1, 8):
                self.apply_style(
                    ws.cell(row=ws.max_row, column=c_idx),
                    is_pk=is_pk and not is_pkfk,
                    is_fk=is_fk and not is_pkfk,
                    is_pkfk=is_pkfk,
                )

        # Anchos de columna
        for letter, width in {"A": 25, "B": 22, "C": 14, "D": 8,
                               "E": 6,  "F": 8,  "G": 45}.items():
            ws.column_dimensions[letter].width = width
        return ws

    # ── Hoja: RELACIONES ENTRE TABLAS ────────────────────────────
    def write_relations_sheet(self, wb, relaciones: list):
        """
        Escribe la hoja 'RELACIONES ENTRE TABLAS' con el formato:
        Tabla Origen | Tabla Destino | Tipo Relación | Descripción
        """
        ws = wb.create_sheet(title="RELACIONES ENTRE TABLAS")

        # Anchos de columna
        for letter, width in {"A": 22, "B": 22, "C": 14, "D": 65}.items():
            ws.column_dimensions[letter].width = width

        # Cabecera
        headers = ["Tabla Origen", "Tabla Destino", "Tipo Relación", "Descripción"]
        ws.append(headers)
        for cell in ws[1]:
            self.apply_style(cell, "header")

        # Filas de datos
        for rel in relaciones:
            ws.append([
                rel.get("tabla_origen", ""),
                rel.get("tabla_destino", ""),
                rel.get("tipo_relacion", ""),
                rel.get("descripcion", ""),
            ])
            for cell in ws[ws.max_row]:
                self.apply_style(cell)

        # Altura de filas de datos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 30

        return ws

    # ── Hoja: ÍNDICES DE LA BASE DE DATOS ────────────────────────
    def write_indexes_sheet(self, wb, indices: list):
        """
        Escribe la hoja 'ÍNDICES DE LA BASE DE DATOS' con el formato:
        Nombre Índice | Tabla | Tipo | Columnas | Único
        """
        ws = wb.create_sheet(title="ÍNDICES DE LA BASE DE DATOS")

        # Anchos de columna
        for letter, width in {"A": 28, "B": 22, "C": 12, "D": 35, "E": 8}.items():
            ws.column_dimensions[letter].width = width

        # Cabecera
        headers = ["Nombre Índice", "Tabla", "Tipo", "Columnas", "Único"]
        ws.append(headers)
        for cell in ws[1]:
            self.apply_style(cell, "header")

        # Filas de datos
        for idx in indices:
            ws.append([
                idx.get("nombre_indice", ""),
                idx.get("tabla", ""),
                idx.get("tipo", "B-Tree"),
                idx.get("columnas", ""),
                idx.get("unico", "No"),
            ])
            for cell in ws[ws.max_row]:
                self.apply_style(cell)

        return ws


# Interfaz gráfica principal de la aplicación
class DataDictApp(ctk.CTk):
    """
    Ventana principal de la aplicación SQL Dictionary Architect.

    Hereda de `ctk.CTk` (CustomTkinter) para proveer la ventana raíz
    con soporte de temas modernos. Orquesta la extracción de metadatos
    (desde BD en vivo o archivo SQL), el enriquecimiento con IA (Groq)
    y la generación del archivo Excel final.

    Atributos principales:
        generator (ExcelGenerator): Instancia responsable de escribir el Excel.
        parser    (SQLFileParser):  Instancia responsable de parsear archivos DDL.
        enricher  (GroqEnricher | None): Instancia del enriquecedor IA, o None si
                                          el usuario no activó esa opción.
    """

    def __init__(self):
        super().__init__()
        self.title("SQL Dictionary Architect")
        self.resizable(True, True)

        # Componentes de lógica
        self.generator = ExcelGenerator()
        self.parser     = SQLFileParser()
        self.enricher   = None

        # Modo de fuente seleccionado por el usuario ("db" o "file")
        self._source_mode = "db"

        # Contenedor con scroll para adaptarse a pantallas pequeñas
        self._scroll = ctk.CTkScrollableFrame(self, label_text="")
        self._scroll.pack(fill="both", expand=True)
        self._scroll.grid_columnconfigure(0, weight=1)
        self._inner = self._scroll

        self._setup_ui()
        self._on_frame_configure()

    def _on_frame_configure(self):
        """Ajusta el tamaño de la ventana al contenido del frame interno."""
        self.update_idletasks()
        w = self._scroll.winfo_reqwidth()  + 40
        h = min(self._scroll.winfo_reqheight() + 60, 900)
        self.geometry(f"{w}x{h}")

    # ── Sector de la interfaz gráfica ────────────────────────────
    def _setup_ui(self):
        p = self._inner

        # ── Título ───────────────────────────────────────────────
        ctk.CTkLabel(p, text="SQL DICTIONARY ARCHITECT",
                     font=("Arial", 24, "bold")).grid(
            row=0, column=0, pady=(28, 4), padx=40)
        ctk.CTkLabel(
            p,
            text="Genera diccionarios de datos desde una BD en vivo o desde archivos SQL / backups",
            font=("Arial", 11), text_color="gray",
        ).grid(row=1, column=0, pady=(0, 16), padx=40)

        # ── Tabs fuente ───────────────────────────────────────────
        tab_frame = ctk.CTkFrame(p, fg_color="transparent")
        tab_frame.grid(row=2, column=0, padx=40, sticky="ew")
        tab_frame.grid_columnconfigure((0, 1), weight=1)

        self.btn_tab_db = ctk.CTkButton(
            tab_frame, text="🗄  Conexión a BD",
            command=lambda: self.switch_source("db"),
            height=38, font=("Arial", 13, "bold"),
        )
        self.btn_tab_file = ctk.CTkButton(
            tab_frame, text="Archivo SQL / Backup",
            command=lambda: self.switch_source("file"),
            height=38, font=("Arial", 13, "bold"),
            fg_color="gray40", hover_color="gray30",
        )
        self.btn_tab_db.grid(row=0, column=0, padx=(0, 4), sticky="ew")
        self.btn_tab_file.grid(row=0, column=1, padx=(4, 0), sticky="ew")

        # ── Panel BD ─────────────────────────────────────────────
        self.panel_db = ctk.CTkFrame(p)
        self.panel_db.grid(row=3, column=0, padx=40, pady=10, sticky="ew")
        self.panel_db.grid_columnconfigure((0, 1), weight=1)
        self._build_db_panel()

        # ── Panel Archivo ─────────────────────────────────────────
        self.panel_file = ctk.CTkFrame(p)
        self.panel_file.grid(row=3, column=0, padx=40, pady=10, sticky="ew")
        self.panel_file.grid_columnconfigure(0, weight=1)
        self._build_file_panel()
        self.panel_file.grid_remove()

        # ── Panel IA ─────────────────────────────────────────────
        self.panel_groq = ctk.CTkFrame(p)
        self.panel_groq.grid(row=4, column=0, padx=40, pady=(4, 4), sticky="ew")
        self.panel_groq.grid_columnconfigure(0, weight=1)
        self._build_groq_panel()

        # ── Barra de progreso ─────────────────────────────────────
        self.progress_var = ctk.DoubleVar(value=0)
        self.progress_bar = ctk.CTkProgressBar(p, variable=self.progress_var, height=10)
        self.progress_bar.grid(row=5, column=0, padx=40, pady=(8, 0), sticky="ew")
        self.progress_bar.grid_remove()

        self.status_label = ctk.CTkLabel(p, text="", font=("Arial", 11), text_color="gray")
        self.status_label.grid(row=6, column=0, pady=(4, 0))

        # ── Botón principal ───────────────────────────────────────
        self.btn_run = ctk.CTkButton(
            p, text="GENERAR DICCIONARIO EXCEL",
            fg_color="#1F3864", hover_color="#15294a",
            font=("Arial", 14, "bold"), height=55,
            command=self.start_process,
        )
        self.btn_run.grid(row=7, column=0, pady=(16, 36), padx=60, sticky="ew")

    # Panel donde va la conexion de BD
    def _build_db_panel(self):
        p = self.panel_db

        ctk.CTkLabel(p, text="Motor de Base de Datos:",
                     font=("Arial", 13, "bold")).grid(
            row=0, column=0, columnspan=2, pady=(15, 5), padx=20, sticky="w")

        self.db_type = ctk.CTkOptionMenu(
            p, values=list(DB_DEFAULTS.keys()), command=self.on_db_change)
        self.db_type.set("PostgreSQL")
        self.db_type.grid(row=1, column=0, columnspan=2, padx=20, pady=5, sticky="ew")

        self.use_defaults_var = ctk.BooleanVar(value=True)
        ctk.CTkSwitch(
            p, text="Usar credenciales por defecto",
            variable=self.use_defaults_var, command=self.toggle_inputs,
        ).grid(row=2, column=0, columnspan=2, pady=12)

        self.host = self._make_input(p, "Host:", "localhost", 3, 0)
        self.port = self._make_input(p, "Puerto:", "5432", 3, 1)
        self.user = self._make_input(p, "Usuario:", "postgres", 5, 0)
        self.pwd = self._make_input(p, "Contraseña:", "", 5, 1, password=True)
        self.db_name = self._make_input(p, "Nombre de la Base de Datos:", "", 7, 0, span=2)
        self.schema = self._make_input(p, "Esquema (Opcional):", "public", 9, 0, span=2)

        self.toggle_inputs()

    def _build_file_panel(self):
        p = self.panel_file

        ctk.CTkLabel(p, text="Archivo SQL o Backup:",
                     font=("Arial", 13, "bold")).grid(
            row=0, column=0, pady=(15, 5), padx=20, sticky="w")

        row = ctk.CTkFrame(p, fg_color="transparent")
        row.grid(row=1, column=0, padx=20, pady=5, sticky="ew")
        row.grid_columnconfigure(0, weight=1)

        self.file_path_entry = ctk.CTkEntry(
            row, height=35,
            placeholder_text="Selecciona un archivo .sql, .backup, .dump…")
        self.file_path_entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ctk.CTkButton(row, text="Examinar", width=110,
                      command=self.browse_sql_file).grid(row=0, column=1)

        ctk.CTkLabel(
            p,
            text="ℹ  Compatible con dumps de PostgreSQL, MySQL, MariaDB y SQL Server.",
            font=("Arial", 10), text_color="gray",
        ).grid(row=2, column=0, padx=20, pady=(4, 14), sticky="w")

    def _build_groq_panel(self):
        p = self.panel_groq

        # Encabezado + switch
        hdr = ctk.CTkFrame(p, fg_color="transparent")
        hdr.grid(row=0, column=0, padx=20, pady=(12, 4), sticky="ew")
        hdr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hdr, text=" Enriquecimiento con IA",
                     font=("Arial", 13, "bold")).grid(row=0, column=0, sticky="w")

        self.use_ai_var = ctk.BooleanVar(value=False)
        ctk.CTkSwitch(
            hdr, text="Activar IA", variable=self.use_ai_var,
            command=self._toggle_groq_inputs,
        ).grid(row=0, column=1, padx=(10, 0))

        # Descripción del proveedor
        self.ai_desc_label = ctk.CTkLabel(
            p,
            text=(
                "Rellena automáticamente con llama-3.3-70b-versatile (Groq):\n"
                "Descripción · Claves Únicas · Observaciones · "
                "Nombre Lógico y Descripción de cada columna.\n"
                "Además genera las hojas de Relaciones e Índices."
            ),
            font=("Arial", 10), text_color="gray", justify="left",
        )
        self.ai_desc_label.grid(row=2, column=0, padx=20, sticky="w")

        # Campo API Key
        key_row = ctk.CTkFrame(p, fg_color="transparent")
        key_row.grid(row=3, column=0, padx=20, pady=(10, 14), sticky="ew")
        key_row.grid_columnconfigure(1, weight=1)

        self.ai_key_label = ctk.CTkLabel(key_row, text="API Key de Groq:",
                                         font=("Arial", 12, "bold"))
        self.ai_key_label.grid(row=0, column=0, padx=(0, 10))

        self.groq_key_entry = ctk.CTkEntry(
            key_row, show="*", placeholder_text="gsk_…", height=35)
        self.groq_key_entry.grid(row=0, column=1, sticky="ew", padx=(0, 8))

        self.btn_show_key = ctk.CTkButton(
            key_row, text="👁", width=36, height=35,
            command=self._toggle_key_visibility)
        self.btn_show_key.grid(row=0, column=2)

        self._key_visible = False
        self._toggle_groq_inputs()  # inicia desactivado

    #  HELPERS DE UI
    def _make_input(self, parent, label: str, default: str,
                    row: int, col: int, password=False, span=1):
        ctk.CTkLabel(parent, text=label,
                     font=("Arial", 12, "bold")).grid(
            row=row, column=col, columnspan=span, padx=20, pady=(10, 0), sticky="w")
        entry = ctk.CTkEntry(parent, show="*" if password else "", height=35)
        entry.insert(0, default)
        entry.grid(row=row + 1, column=col, columnspan=span,
                   padx=20, pady=(0, 10), sticky="ew")
        return entry

    def _toggle_groq_inputs(self):
        state = "normal" if self.use_ai_var.get() else "disabled"
        self.groq_key_entry.configure(state=state)
        self.btn_show_key.configure(state=state)

    def _toggle_key_visibility(self):
        self._key_visible = not self._key_visible
        self.groq_key_entry.configure(show="" if self._key_visible else "*")

    def on_db_change(self, choice=None):
        choice = choice or self.db_type.get()
        if self.use_defaults_var.get():
            defaults = DB_DEFAULTS.get(choice, {})
            for field, key in [(self.port, "port"), (self.user, "user")]:
                field.configure(state="normal")
                field.delete(0, "end")
                field.insert(0, defaults.get(key, ""))
                field.configure(state="disabled")
        else:
            for field in (self.port, self.user):
                field.configure(state="normal")

    def toggle_inputs(self):
        self.on_db_change(self.db_type.get())

    def switch_source(self, mode):
        self._source_mode = mode
        if mode == "db":
            self.btn_tab_db.configure(fg_color="#1F3864", hover_color="#15294a")
            self.btn_tab_file.configure(fg_color="gray40", hover_color="gray30")
            self.panel_db.grid()
            self.panel_file.grid_remove()
        else:
            self.btn_tab_file.configure(fg_color="#1F3864", hover_color="#15294a")
            self.btn_tab_db.configure(fg_color="gray40", hover_color="gray30")
            self.panel_file.grid()
            self.panel_db.grid_remove()
        self.after(50, self._on_frame_configure)

    def browse_sql_file(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo SQL o Backup",
            filetypes=[
                ("Archivos SQL", "*.sql"),
                ("Backups", "*.backup *.dump *.bak"),
                ("Todos los archivos", "*.*"),
            ],
        )
        if path:
            self.file_path_entry.delete(0, "end")
            self.file_path_entry.insert(0, path)

    def set_status(self, msg: str, progress: float | None = None):
        """
        Actualiza el label de estado y la barra de progreso en la UI.

        Args:
            msg      : Texto descriptivo del paso actual.
            progress : Valor entre 0.0 y 1.0 para la barra de progreso.
                       Si es None, no se modifica la barra.
        """
        self.status_label.configure(text=msg)
        if progress is not None:
            self.progress_var.set(progress)
        self.update_idletasks()

    # ── Proceso principal ────────────────────────────────────────
    def start_process(self):
        """
        Punto de entrada del botón principal. Deshabilita el botón,
        muestra la barra de progreso y lanza `_run` en un hilo secundario
        para no bloquear la interfaz gráfica.
        """
        self.btn_run.configure(state="disabled", text="⏳  Procesando…")
        self.progress_bar.grid()
        self.progress_var.set(0)
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        """
        Flujo principal ejecutado en hilo secundario:
        1. Inicializa el enriquecedor IA si fue activado.
        2. Extrae metadatos desde la BD o el archivo SQL.
        3. Enriquece cada tabla con la IA (nombre lógico, descripción, etc.).
        4. Genera la hoja de relaciones e índices en una llamada final a la IA.
        5. Abre el diálogo para guardar el archivo Excel resultante.
        """
        try:
            # ── Inicializar enriquecedor ──────────────────────
            self.enricher = None
            if self.use_ai_var.get():
                key = self.groq_key_entry.get().strip()
                if not key:
                    messagebox.showwarning(
                        "API Key requerida",
                        "Ingresa tu API Key para usar el enriquecimiento con IA.")
                    return
                try:
                    self.enricher = GroqEnricher(key)
                except Exception as e:
                    messagebox.showerror("Error al inicializar Groq", str(e))
                    return

            # ── Extracción ────────────────────────────────────
            if self._source_mode == "db":
                tables_data = self._extract_from_db()
            else:
                tables_data = self._extract_from_file()

            if tables_data is None:
                return

            # ── Enriquecimiento IA ────────────────────────────
            extra_data = {"relaciones": [], "indices": []}
            if self.enricher:
                total = len(tables_data)
                for i, (t_name, info) in enumerate(tables_data.items()):
                    self.set_status(
                        f"  Enriqueciendo con IA  {i + 1}/{total}: {t_name}",
                        0.5 + (i + 1) / total * 0.40,
                    )
                    try:
                        info["enriched"] = self.enricher.enrich_table(
                            t_name,
                            info["columns"],
                            info["pk"],
                            info["fk_map"],
                            info.get("schema"),
                        )
                    except Exception as exc:
                        info["enriched"] = {}
                        print(f"[IA] Error en tabla '{t_name}': {exc}")

                # Generar relaciones e índices con una sola llamada al final
                self.set_status("Generando relaciones e índices…", 0.93)
                try:
                    extra_data = self.enricher.generate_relations_and_indexes(tables_data)
                except Exception as exc:
                    print(f"[IA] Error generando relaciones/índices: {exc}")

            self._save_excel(tables_data, extra_data)

        except Exception as e:
            messagebox.showerror("Error inesperado", str(e))
        finally:
            self.btn_run.configure(state="normal", text="GENERAR DICCIONARIO EXCEL")
            self.progress_bar.grid_remove()
            self.status_label.configure(text="")

    # ── Extracción desde motor de BD ─────────────────────────────
    def _extract_from_db(self):
        """
        Conecta al motor de BD configurado en la UI y extrae los metadatos
        de todas las tablas del esquema indicado usando SQLAlchemy Inspector.

        Incluye un fallback para MySQL/MariaDB que consulta INFORMATION_SCHEMA
        cuando el inspector no detecta las FK correctamente.

        Returns:
            dict: {nombre_tabla: {columns, pk, fk_map, schema, enriched}} o None si falla.
        """
        db_t = self.db_type.get()
        config = DB_DEFAULTS[db_t]

        if db_t == "SQLite":
            conn_str = f"sqlite:///{self.db_name.get()}"
        else:
            conn_str = (
                f"{config['driver']}://{self.user.get()}:"
                f"{self.pwd.get()}@{self.host.get()}:"
                f"{self.port.get()}/{self.db_name.get()}"
            )
        if db_t == "Microsoft SQL Server":
            conn_str += "?driver=ODBC+Driver+17+for+SQL+Server"

        self.set_status("Conectando a la base de datos…", 0.05)
        engine = create_engine(conn_str)
        inspector = inspect(engine)
        schema = self.schema.get() or None
        t_names = inspector.get_table_names(schema=schema)

        tables_data = {}
        total = len(t_names)

        for i, t_name in enumerate(t_names):
            self.set_status(f"Procesando tabla {i + 1}/{total}: {t_name}",
                            (i + 1) / total * 0.45)

            pk_cols = inspector.get_pk_constraint(
                t_name, schema=schema).get("constrained_columns", [])

            # FK map robusto
            fk_map = {}
            for fk in inspector.get_foreign_keys(t_name, schema=schema):
                srcs = fk.get("constrained_columns", [])
                reft = fk.get("referred_table", "")
                refc = fk.get("referred_columns", [])
                for sc, rc in zip(srcs, refc):
                    fk_map[sc] = f"{reft}.{rc}"

            # Fallback MySQL/MariaDB via INFORMATION_SCHEMA
            if not fk_map and db_t in ("MySQL", "MariaDB"):
                try:
                    with engine.connect() as conn:
                        res = conn.execute(text("""
                            SELECT COLUMN_NAME, REFERENCED_TABLE_NAME,
                                   REFERENCED_COLUMN_NAME
                            FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE
                            WHERE TABLE_SCHEMA = :schema
                              AND TABLE_NAME   = :table
                              AND REFERENCED_TABLE_NAME IS NOT NULL
                        """), {"schema": self.db_name.get(), "table": t_name})
                        for row in res:
                            fk_map[row[0]] = f"{row[1]}.{row[2]}"
                except Exception:
                    pass

            # Columnas
            columns = []
            for col in inspector.get_columns(t_name, schema=schema):
                t_obj = col["type"]
                t_str = str(t_obj).split("(")[0].upper()
                length = (str(getattr(t_obj, "length", "—"))
                          if getattr(t_obj, "length", None) else "—")
                columns.append({
                    "name": col["name"],
                    "type": t_str,
                    "length": length,
                    "nullable": col.get("nullable", True),
                })

            tables_data[t_name] = {
                "columns": columns,
                "pk": pk_cols,
                "fk_map": fk_map,
                "schema": schema,
                "enriched": None,
            }

        return tables_data

    # ── Extracción desde archivo SQL / backup ────────────────────
    def _extract_from_file(self):
        """
        Lee el archivo SQL seleccionado por el usuario y extrae la estructura
        DDL usando `SQLFileParser`.

        Returns:
            dict: {nombre_tabla: {columns, pk, fk_map, schema, enriched}} o None si falla.
        """
        filepath = self.file_path_entry.get().strip()
        if not filepath:
            messagebox.showwarning("Sin archivo",
                                   "Por favor selecciona un archivo SQL o backup.")
            return None

        self.set_status("Leyendo y parseando archivo SQL…", 0.1)
        try:
            parsed = self.parser.parse(filepath)
        except Exception as e:
            messagebox.showerror("Error al leer archivo", str(e))
            return None

        if not parsed:
            messagebox.showwarning(
                "Sin tablas detectadas",
                "No se encontraron sentencias CREATE TABLE en el archivo.\n"
                "Asegúrate de que sea un script DDL o un backup con estructura.")
            return None

        total = len(parsed)
        tables_data = {}
        for i, (t_name, info) in enumerate(parsed.items()):
            self.set_status(f"Procesando tabla {i + 1}/{total}: {t_name}",
                            (i + 1) / total * 0.45)
            tables_data[t_name] = {
                "columns": info["columns"],
                "pk": info["pk"],
                "fk_map": info["fk_map"],
                "schema": None,
                "enriched": None,
            }
        return tables_data

    # ── Guardado del archivo Excel ────────────────────────────────
    def _save_excel(self, tables_data: dict, extra_data: dict | None = None):
        """
        Abre el diálogo "Guardar como", crea el Workbook de openpyxl y escribe:
          - Una hoja por tabla (mediante ExcelGenerator.write_table_sheet).
          - Una hoja ÍNDICE al inicio con el listado y descripción de todas las tablas.
          - Una hoja de relaciones entre tablas.
          - Una hoja de índices recomendados.

        Args:
            tables_data : Resultado del paso de extracción/enriquecimiento.
            extra_data  : Dict {"relaciones": [...], "indices": [...]} generado por la IA.
        """
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar diccionario de datos",
        )
        if not path:
            return

        self.set_status("Generando archivo Excel…", 0.97)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for t_name, info in tables_data.items():
            self.generator.write_table_sheet(
                wb, t_name,
                info["columns"], info["pk"], info["fk_map"],
                info.get("schema"), info.get("enriched"),
            )

        # Hoja de índice (con descripción IA si existe)
        ws_idx = wb.create_sheet(title="ÍNDICE", index=0)
        for col_letter, w in zip("ABCDE", [7, 35, 40, 18, 25]):
            ws_idx.column_dimensions[col_letter].width = w

        headers = ["N.°", "Nombre Tabla", "Descripción", "Esquema", "Módulo del Sistema"]
        ws_idx.append(headers)
        for cell in ws_idx[1]:
            self.generator.apply_style(cell, "header")

        for idx, (t_name, info) in enumerate(tables_data.items(), start=1):
            ai_desc = (info.get("enriched") or {}).get("descripcion", "")
            ws_idx.append([idx, t_name, ai_desc, info.get("schema") or "", ""])
            for cell in ws_idx[idx + 1]:
                self.generator.apply_style(cell)

        # Hojas de relaciones e índices (siempre al final)
        ed = extra_data or {}
        self.generator.write_relations_sheet(wb, ed.get("relaciones", []))
        self.generator.write_indexes_sheet(wb, ed.get("indices", []))

        wb.save(path)
        n = len(tables_data)
        messagebox.showinfo(
            "Éxito",
            f"Diccionario generado correctamente.\n{n} tabla(s) procesada(s).\n\n{path}",
        )


# ── Punto de entrada ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    DataDictApp().mainloop()